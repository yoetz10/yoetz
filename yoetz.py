import asyncio
import os
import time
from datetime import datetime
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from email.mime.text import MIMEText
import base64
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from telegram import Update
from telegram.ext import Application, MessageHandler, filters, ContextTypes
import email
import openpyxl
from dotenv import load_dotenv
import re

# טעינת משתני סביבה
load_dotenv()
print(f"BOT_TOKEN: {os.getenv('BOT_TOKEN')}")
print(f"GOOGLE_CREDENTIALS_PATH: {os.getenv('GOOGLE_CREDENTIALS_PATH')}")

SCOPES = [
    'https://www.googleapis.com/auth/gmail.send',
    'https://www.googleapis.com/auth/gmail.readonly',
    'https://www.googleapis.com/auth/gmail.modify'
]

CREDENTIALS_FILE = os.getenv('GOOGLE_CREDENTIALS_PATH', 'credentials/credentials.json')
TOKEN_FILE = 'token.json'
QUESTIONS_FILE = "questions.xlsx"

# מבני נתונים גלובליים
user_questions = {}

# הגדרת המומחים עם פרטים נוספים
EXPERTS = {
    "yorvada@gmail.com": {"name": "יורם", "title": "יועץ משפחתי"},
    "nomitrapi@gmail.com": {"name": "נעמי", "title": "פסיכולוגית"},
    "phdelazar@gmail.com": {"name": "פיליפ", "title": "יועץ זוגי"}
}

def load_questions_from_excel():
    """טוען שאלות מהאקסל למבנה הנתונים בזיכרון"""
    try:
        if not os.path.exists(QUESTIONS_FILE):
            return
        
        workbook = openpyxl.load_workbook(QUESTIONS_FILE)
        sheet = workbook.active
        
        for row in sheet.iter_rows(min_row=2):  # מתחיל משורה 2 כדי לדלג על הכותרות
            if row[0].value:  # אם יש מזהה שאלה
                question_id = str(row[0].value)
                user_questions[question_id] = {
                    "question": row[1].value,
                    "user_name": row[3].value,
                    "chat_id": row[4].value,
                    "timestamp": datetime.now().isoformat()
                }
        print(f"Loaded {len(user_questions)} questions from Excel")
    except Exception as e:
        print(f"Error loading questions from Excel: {e}")

def decode_email_subject(subject):
    """מפענח את נושא המייל מקידוד UTF-8"""
    try:
        if not subject:
            return ""
        
        # אם המייל מקודד ב-UTF-8
        if "=?UTF-8?" in subject:
            # מנקה רווחים מיותרים וסימני שורה חדשה
            subject = subject.replace('\n', '').replace('\r', '').strip()
            # מפענח את הקידוד באמצעות חבילת email
            decoded_parts = email.header.decode_header(subject)
            decoded_subject = ''
            for part, charset in decoded_parts:
                if isinstance(part, bytes):
                    try:
                        decoded_subject += part.decode(charset or 'utf-8')
                    except:
                        decoded_subject += part.decode('utf-8', errors='replace')
                else:
                    decoded_subject += str(part)
            return decoded_subject
        return subject
    except Exception as e:
        print(f"Error decoding subject: {e}")
        return subject

def get_email_content(mime_msg):
    """מחלץ את תוכן המייל מאובייקט MIME"""
    try:
        if mime_msg.is_multipart():
            for part in mime_msg.walk():
                if part.get_content_type() == 'text/plain':
                    content = part.get_payload(decode=True)
                    try:
                        return content.decode('utf-8')
                    except UnicodeDecodeError:
                        return content.decode('iso-8859-1')
        else:
            content = mime_msg.get_payload(decode=True)
            try:
                return content.decode('utf-8')
            except UnicodeDecodeError:
                return content.decode('iso-8859-1')
    except Exception as e:
        print(f"Error getting email content: {e}")
        return None

def create_message(sender, to, subject, body):
    """יוצר הודעת מייל"""
    try:
        message = MIMEText(body, 'plain', 'utf-8')
        message['to'] = to
        message['from'] = sender
        message['subject'] = subject
        return {'raw': base64.urlsafe_b64encode(message.as_bytes()).decode()}
    except Exception as e:
        print(f"Error creating message: {e}")
        return None

def send_message(service, sender, message):
    """שולח הודעת מייל"""
    try:
        if not message:
            print("Message is None, cannot send.")
            return None
        return service.users().messages().send(userId=sender, body=message).execute()
    except Exception as error:
        print(f"Error sending message: {error}")
        return None

def authenticate_gmail_api():
    """מאמת מול Gmail API"""
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception:
                os.remove(TOKEN_FILE)
                creds = None

        if not creds:
            if not os.path.exists(CREDENTIALS_FILE):
                raise FileNotFoundError(f"Missing credentials file at {CREDENTIALS_FILE}")
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
            with open(TOKEN_FILE, 'w') as token:
                token.write(creds.to_json())

    return build('gmail', 'v1', credentials=creds)

def ensure_excel_file_exists():
    """מוודא שקובץ האקסל קיים"""
    if not os.path.exists(QUESTIONS_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["ID", "Question", "Answer", "User Name", "Chat ID", "Expert"])
        workbook.save(QUESTIONS_FILE)

def save_question_to_excel(question_id, question_text, user_name, chat_id, answer=None, expert=None):
    """שומר שאלה או תשובה לקובץ אקסל"""
    try:
        workbook = openpyxl.load_workbook(QUESTIONS_FILE)
        sheet = workbook.active
        
        # בדיקה אם השאלה כבר קיימת
        question_row = None
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if str(row[0].value) == str(question_id):
                question_row = row_idx
                break

        if question_row:
            # עדכון שורה קיימת
            sheet.cell(row=question_row, column=3, value=answer)  # עמודת תשובה
            sheet.cell(row=question_row, column=6, value=expert)  # עמודת מומחה
        else:
            # הוספת שורה חדשה
            sheet.append([question_id, question_text, answer, user_name, chat_id, expert])

        workbook.save(QUESTIONS_FILE)
        print(f"Successfully saved/updated question {question_id} in Excel")
    except Exception as e:
        print(f"Error saving to Excel: {e}")

def clean_expert_response(content):
    """ניקוי תשובת המומחה מתוכן מיותר"""
    if not content:
        return ""

    lines = content.split('\n')
    cleaned_lines = []
    
    skip_markers = [
        'original message',
        'מהמייל המקורי',
        'שאלה חדשה התקבלה',
        'מזהה שאלה:',
        'שואל:',
        'שאלה:',
        '------',
        'From:',
        'Sent:',
        'To:',
        'Subject:'
    ]
    
    for line in lines:
        line = line.strip()
        if (
            line and
            not line.startswith('>') and
            not line.startswith('On') and
            not any(marker.lower() in line.lower() for marker in skip_markers)
        ):
            cleaned_lines.append(line)
    
    return '\n'.join(cleaned_lines).strip()

async def handle_question(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """מטפל בשאלה חדשה מהמשתמש"""
    try:
        question_text = update.message.text
        user_name = update.message.from_user.full_name
        chat_id = update.message.chat_id
        
        # יצירת מזהה ייחודי לשאלה
        question_id = str(int(time.time()))
        
        # שמירת פרטי השאלה במבנה הנתונים
        user_questions[question_id] = {
            "user_name": user_name,
            "chat_id": chat_id,
            "question": question_text,
            "timestamp": datetime.now().isoformat()
        }

        # שמירה באקסל
        save_question_to_excel(question_id, question_text, user_name, chat_id)

        # שליחת השאלה למומחים
        service = authenticate_gmail_api()
        question_subject = f"שאלה חדשה #{question_id} מאת {user_name}"
        question_body = f"""
שאלה חדשה התקבלה:

מזהה שאלה: {question_id}
שואל: {user_name}
שאלה:
{question_text}

אנא השב למייל זה (Reply) כדי שהתשובה תגיע לשואל.
"""

        for expert_email in EXPERTS.keys():
            message = create_message(
                "yoetz10bot@gmail.com",
                expert_email,
                question_subject,
                question_body
            )
            if message:
                send_message(service, "yoetz10bot@gmail.com", message)

        await update.message.reply_text("✅ השאלה נשלחה למומחים. תקבל תשובה בקרוב.")

    except Exception as e:
        print(f"Error in handle_question: {e}")
        await update.message.reply_text("❌ אירעה שגיאה בעת שליחת השאלה. אנא נסה שוב.")

async def check_for_answers(context: ContextTypes.DEFAULT_TYPE):
    """בודק תשובות חדשות מהמומחים"""
    try:
        print("Checking for answers...")
        service = authenticate_gmail_api()
        
        # חיפוש תשובות שלא נקראו
        results = service.users().messages().list(
            userId='me',
            q="is:unread"  # מחפש את כל ההודעות שלא נקראו
        ).execute()
        messages = results.get('messages', [])

        if not messages:
            return

        for message in messages:
            try:
                msg = service.users().messages().get(userId='me', id=message['id'], format='raw').execute()
                msg_str = base64.urlsafe_b64decode(msg['raw'].encode('ASCII'))
                mime_msg = email.message_from_bytes(msg_str)

                # חילוץ וקידוד פרטי המייל
                subject = decode_email_subject(mime_msg['Subject'] or "")
                sender = mime_msg['From']
                
                print(f"Processing email from {sender} with subject: {subject}")

                # חילוץ כתובת המייל של השולח
                if '<' in sender:
                    sender = sender.split('<')[1].strip('>')
                
                # וידוא שהשולח הוא מומחה מורשה
                if sender.lower() not in EXPERTS:
                    print(f"Unauthorized response from {sender}")
                    continue

                # חילוץ מזהה השאלה מהנושא
                try:
                    # מחפש מספר אחרי הסימן #
                    question_id_match = re.search(r'#(\d+)', subject)
                    if not question_id_match:
                        print(f"No question ID found in subject: {subject}")
                        continue
                    
                    question_id = question_id_match.group(1)
                    print(f"Found question ID: {question_id}")
                except Exception as e:
                    print(f"Error extracting question ID from subject '{subject}': {e}")
                    continue

                # וידוא שהשאלה קיימת במערכת
                if question_id not in user_questions:
                    print(f"Question ID {question_id} not found in active questions")
                    print(f"Available question IDs: {list(user_questions.keys())}")
                    # נסה לטעון מחדש מהאקסל
                    load_questions_from_excel()
                    if question_id not in user_questions:
                        print(f"Question still not found after reloading from Excel")
                        continue

                # חילוץ וניקוי התשובה
                answer_content = get_email_content(mime_msg)
                if not answer_content or len(answer_content.strip()) < 10:
                    print(f"Empty or too short answer from {sender}")
                    continue

                # ניקוי התשובה
                cleaned_answer = clean_expert_response(answer_content)
                print(f"Cleaned answer length: {len(cleaned_answer)}")
                
                # שליחת התשובה למשתמש
                try:
                    chat_id = user_questions[question_id]["chat_id"]
                    original_question = user_questions[question_id]["question"]
                    expert_info = EXPERTS[sender.lower()]

                    response_text = f"""
✨ קיבלת תשובה ממומחה!

👨‍⚕️ המשיב/ה: {expert_info['name']} - {expert_info['title']}

📝 השאלה המקורית:
{original_question}

✍️ התשובה:
{cleaned_answer}

תודה שהשתמשת במערכת הייעוץ שלנו! 🙏
"""
                    await context.bot.send_message(
                        chat_id=chat_id,
                        text=response_text
                    )

                    # עדכון בקובץ Excel
                    save_question_to_excel(
                        question_id,
                        original_question,
                        user_questions[question_id]["user_name"],
                        chat_id,
                        cleaned_answer,
                        f"{expert_info['name']} ({expert_info['title']})"
                    )

                    print(f"Successfully delivered answer from {expert_info['name']} for question {question_id}")

                except Exception as e:
                    print(f"Error sending answer to user: {e}")
                    continue

            finally:
                # סימון המייל כנקרא
                try:
                    service.users().messages().modify(
                        userId='me',
                        id=message['id'],
                        body={'removeLabelIds': ['UNREAD']}
                    ).execute()
                except Exception as e:
                    print(f"Error marking message as read: {e}")

    except Exception as e:
        print(f"Error in check_for_answers: {e}")

async def main():
    """פונקציית הפעלה ראשית"""
    try:
        BOT_TOKEN = os.getenv("BOT_TOKEN")
        if not BOT_TOKEN:
            raise ValueError("Missing BOT_TOKEN")

        ensure_excel_file_exists()
        
        application = Application.builder().token(BOT_TOKEN).build()
        application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_question))
        application.job_queue.run_repeating(check_for_answers, interval=60)  # בדיקה כל דקה

        print("Starting bot...")
        await application.initialize()
        await application.start()
        await application.updater.start_polling()

        while True:
            await asyncio.sleep(1)

    except Exception as e:
        print(f"Critical error in main: {e}")
    finally:
        if 'application' in locals():
            await application.stop()
            await application.shutdown()

if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nBot stopped by user")
    except Exception as e:
        print(f"Critical error: {e}")